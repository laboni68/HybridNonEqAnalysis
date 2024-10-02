from collections import defaultdict
import pandas as pd
import sys
import os
import openpyxl
import xlsxwriter

dir = "../"

benchmarks = ["benchmarksHumaneval", "QuixBugs"]
subdirs = ["Eq", "NEq"]
resultFiles = ["BFuzz2000.xlsx","BFuzz3000.xlsx","PrePostFuzzD1097Changed2000.xlsx","BFuzz5000.xlsx","PrePostFuzzD10115Changed2000.xlsx"]

out_dir = dir + "/" + benchmarks[0] + "/"

for resultFile in resultFiles:
    outFile = out_dir + "Total" + resultFile
    outWorkbook = xlsxwriter.Workbook(outFile)
    outWorksheet = outWorkbook.add_worksheet()
    outWorksheet.write("A1", "File Name")
    outWorksheet.write("B1", "Result")
    outWorksheet.write("C1", "Time (ms)")
    outWorksheet.write("D1", "Total Equivalent")
    outWorksheet.write("E1", "Total Not Equivalent")
    outWorksheet.write("F1", "Total Unknown")
    outWorksheet.write("G1", "Total time (ms)")
    outWorksheet.write("H1", "Avg time (s)")
    outWorksheet.write("I1", "Avg Eq time (s)")
    outWorksheet.write("J1", "Avg NEq time (s)")
    outWorksheet.write("K1", "Avg Unknown time (s)")
    cur_row = 2
    eq_count, neq_count, unknown_count, total_time, eq_time, neq_time, unk_time = 0, 0, 0, 0, 0, 0, 0  
    for benchmark in benchmarks:
        cur_benchmark = dir + "/" + benchmark + "/"
        for subdir in subdirs:
            cur_dir = cur_benchmark + subdir + "/"
            inFile = cur_dir + resultFile
            inWorkbook = openpyxl.load_workbook(inFile)
            inWorksheet = inWorkbook['Sheet1']
            for row in inWorksheet.iter_rows(min_row=2,max_row=inWorksheet.max_row,values_only=True):
                # print(row)
                outWorksheet.write("A"+str(cur_row), row[0])
                outWorksheet.write("B"+str(cur_row), row[1])
                outWorksheet.write("C"+str(cur_row), row[2])
                if row[1] == "EQUIVALENT" or row[1] == "EQ":
                    eq_count += 1
                    eq_time += float(row[2])
                elif row[1] == "NOT EQUIVALENT" or row[1] == "NEQ":
                    neq_count += 1
                    neq_time += float(row[2])
                elif row[1] == "UNKNOWN" or row[1] == None or row[1] == "MAYBE_EQ" or row[1] == "Null" or row[1] == "ERROR":
                    unknown_count += 1
                    unk_time += float(row[2])
                total_time += float(row[2])
                cur_row += 1
    outWorksheet.write("D2", eq_count)
    outWorksheet.write("E2", neq_count)
    outWorksheet.write("F2", unknown_count)
    outWorksheet.write("G2", total_time)
    outWorksheet.write("H2", total_time/(1000*(cur_row - 2)))
    if eq_count != 0:
        outWorksheet.write("I2", eq_time/(1000*eq_count))
    if neq_count != 0:
        outWorksheet.write("J2", neq_time/(1000*neq_count))
    if unknown_count != 0:
        outWorksheet.write("K2", unk_time/(1000*unknown_count))
    outWorkbook.close()
