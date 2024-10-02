from collections import defaultdict
import pandas as pd
import sys
import os
import openpyxl
import xlsxwriter

benchmark = sys.argv[1]

# dir = "/media/laboni/HDD11/PASDA/benchmarks/"
# dir = "/media/laboni/HDD11/PASDA/EqBench/"
dir = "/media/laboni/HDD11/PASDA/"+benchmark+"/"

outFileUniform = dir + "uniformFuzz.xlsx"
outFileBoundary = dir + "boundaryFuzz.xlsx"
outFileCombined = dir + "combinedFuzz.xlsx"

times = ["500","1000", "2000", "3000", "4000", "5000", "6000","7000","8000","9000","10000"]
iters = ["1", "2", "3", "4", "5"]

iter2resultMap = {"1":"B", "2":"C", "3":"D", "4":"E", "5":"F"}
iter2timeMap = {"1":"H", "2":"I", "3":"J", "4":"K", "5":"L"}

if benchmark == "benchmarks":
    totalNeq=73 
    total=142
elif benchmark == "EqBench":
    totalNeq=23
    total=50
# totalNeq=73 
# total=142
# totalNeq=21
# total=51

start, end = 2, total+1
# start, end = 2, 125

row=2
outCombinedWorkbook = xlsxwriter.Workbook(outFileCombined)
outCombinedWorksheetPercent = outCombinedWorkbook.add_worksheet("percent")
outCombinedWorksheetTotal = outCombinedWorkbook.add_worksheet("total")
outCombinedWorksheetPercent.write('A1', "Time (ms)")
outCombinedWorksheetPercent.write('B1', "Uniform")
outCombinedWorksheetPercent.write('C1', "Boundary")
outCombinedWorksheetTotal.write('A1', "Time (ms)")
outCombinedWorksheetTotal.write('B1', "Uniform")
outCombinedWorksheetTotal.write('C1', "Boundary")

for time in times:
    outCombinedWorksheetPercent.write('A'+str(row), time)
    outCombinedWorksheetTotal.write('A'+str(row), time)
    row += 1
 

uniformResultMap = dict()
boundaryResultMap = dict()
row2folderMap = dict()

row=2
outWorkbook = xlsxwriter.Workbook(outFileUniform)
for time in times:
    outWorksheet = outWorkbook.add_worksheet(time)
    outWorksheet.write('A1', "Folder")
    outWorksheet.write('B1', "Result 1")
    outWorksheet.write('C1', "Result 2")
    outWorksheet.write('D1', "Result 3")
    outWorksheet.write('E1', "Result 4")
    outWorksheet.write('F1', "Result 5")
    outWorksheet.write('G1', "Best")
    outWorksheet.write('H1', "Time 1")
    outWorksheet.write('I1', "Time 2")
    outWorksheet.write('J1', "Time 3")
    outWorksheet.write('K1', "Time 4")
    outWorksheet.write('L1', "Time 5")
    results = []
    for iter in iters:
        inWorkbook = openpyxl.load_workbook(dir+"1Fuzz"+time+iter+".xlsx")
        inWorksheet = inWorkbook['Sheet1']
        if iter == "1":
            for k in range(start, end+1):
                folderName = inWorksheet['A'+str(k)].value
                outWorksheet.write('A'+str(k), folderName)
                uniformResultMap[folderName] = defaultdict(int)
                row2folderMap[k] = folderName
        resCol = iter2resultMap[iter]
        timeCol = iter2timeMap[iter]
        resCount = 0
        for k in range(start, end+1):
            result = inWorksheet['B'+str(k)].value
            outWorksheet.write(resCol+str(k), result)
            outWorksheet.write(timeCol+str(k), inWorksheet['C'+str(k)].value)
            uniformResultMap[inWorksheet['A'+str(k)].value][result] += 1
            if result == "NEq":
                resCount += 1
        outWorksheet.write(resCol+str(end+1), str(resCount))
        results.append(resCount)
    outWorksheet.write('A'+str(end+1), str(sum(results)/len(results)))
    outWorksheet.write('M'+str(end+1), str(sum(results)))

    outCombinedWorksheetPercent.write('B'+str(row), (sum(results)*100)/(totalNeq*len(iters)))
    outCombinedWorksheetTotal.write('B'+str(row), sum(results))
    row += 1
    count = 0
    for k in range(start, end+1):
        if uniformResultMap[row2folderMap[k]]["NEq"] >= 3:
            outWorksheet.write('G'+str(k), "1")
            count += 1
    outWorksheet.write('G'+str(end+1), str(count))
outWorkbook.close()      

row2folderMap = dict()
row=2
outWorkbook = xlsxwriter.Workbook(outFileBoundary)
for time in times:
    outWorksheet = outWorkbook.add_worksheet(time)
    outWorksheet.write('A1', "Folder")
    outWorksheet.write('B1', "Result 1")
    outWorksheet.write('C1', "Result 2")
    outWorksheet.write('D1', "Result 3")
    outWorksheet.write('E1', "Result 4")
    outWorksheet.write('F1', "Result 5")
    outWorksheet.write('G1', "Best")
    outWorksheet.write('H1', "Time 1")
    outWorksheet.write('I1', "Time 2")
    outWorksheet.write('J1', "Time 3")
    outWorksheet.write('K1', "Time 4")
    outWorksheet.write('L1', "Time 5")
    results = []
    for iter in iters:
        inWorkbook = openpyxl.load_workbook(dir+"1BFuzz"+time+iter+".xlsx")
        inWorksheet = inWorkbook['Sheet1']
        if iter == "1":
            for k in range(start, end+1):
                folderName = inWorksheet['A'+str(k)].value
                outWorksheet.write('A'+str(k), folderName)
                boundaryResultMap[folderName] = defaultdict(int)
                row2folderMap[k] = folderName
        resCol = iter2resultMap[iter]
        timeCol = iter2timeMap[iter]
        resCount = 0
        for k in range(start, end+1):
            result = inWorksheet['B'+str(k)].value
            outWorksheet.write(resCol+str(k), result)
            outWorksheet.write(timeCol+str(k), inWorksheet['C'+str(k)].value)
            boundaryResultMap[inWorksheet['A'+str(k)].value][result] += 1
            if result == "NEq":
                resCount += 1
        outWorksheet.write(resCol+str(end+1), str(resCount))
        results.append(resCount)
    outWorksheet.write('A'+str(end+1), str(sum(results)/len(results)))
    outWorksheet.write('M'+str(end+1), str(sum(results)))
    outCombinedWorksheetPercent.write('C'+str(row), (sum(results)*100)/(totalNeq*len(iters)))
    outCombinedWorksheetTotal.write('C'+str(row), sum(results))
    # outCombinedWorksheet.write('C'+str(row), (sum(results)*100)/(totalNeq*len(iters)))
    row += 1    
    count = 0
    for k in range(start, end+1):
        if boundaryResultMap[row2folderMap[k]]["NEq"] >= 3:
            outWorksheet.write('G'+str(k), "1")
            count += 1
    outWorksheet.write('G'+str(end+1), str(count))
outWorkbook.close()   
outCombinedWorkbook.close()
# os.system("bash removeExcel.sh")