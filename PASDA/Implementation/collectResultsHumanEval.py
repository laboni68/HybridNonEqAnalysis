from collections import defaultdict
import pandas as pd
import sys
import os
import openpyxl
import xlsxwriter

dir = "../"

benchmarks = ["benchmarksHumaneval", "QuixBugs"]
subdirs = ["Eq", "NEq"]
resultFiles = ["ResultPASDA120.xlsx"]

out_dir = dir + "/" + benchmarks[0] + "/"

for resultFile in resultFiles:
    outFile = out_dir + "Total" + resultFile
    outWorkbook = xlsxwriter.Workbook(outFile)
    outWorksheet = outWorkbook.add_worksheet()
    outWorksheet.write("A1", "File Name")
    outWorksheet.write("B1", "Result")
    outWorksheet.write("C1", "Time (ms)")
    outWorksheet.write("D1", "Total Equivalent")
    outWorksheet.write("E1", "Total Not Equivalent")
    outWorksheet.write("F1", "Total Unknown")
    outWorksheet.write("G1", "Total time (ms)")
    outWorksheet.write("H1", "Avg time (s)")
    outWorksheet.write("I1", "Avg Eq time (s)")
    outWorksheet.write("J1", "Avg NEq time (s)")
    outWorksheet.write("K1", "Avg Unknown time (s)")
    outWorksheet.write("L1", "Total maybe_eq")
    outWorksheet.write("M1", "Total maybe_neq")
    outWorksheet.write("N1", "Avg maybe_eq time (s)")
    outWorksheet.write("O1", "Avg maybe_neq time (s)")
    cur_row = 2
    eq_count, neq_count, unknown_count, total_time, eq_time, neq_time, unk_time = 0, 0, 0, 0, 0, 0, 0  
    maybe_eq_count, maybe_neq_count, maybe_eq_time, maybe_neq_time = 0, 0, 0, 0
    for benchmark in benchmarks:
        cur_benchmark = dir + "/" + benchmark + "/"
        for subdir in subdirs:
            cur_dir = cur_benchmark + subdir + "/"
            inFile = cur_dir + resultFile
            inWorkbook = openpyxl.load_workbook(inFile)
            inWorksheet = inWorkbook['Sheet1']
            for row in inWorksheet.iter_rows(min_row=2,max_row=inWorksheet.max_row,values_only=True):
                # print(row)
                outWorksheet.write("A"+str(cur_row), row[0])
                outWorksheet.write("B"+str(cur_row), row[1])
                outWorksheet.write("C"+str(cur_row), row[2])
                if row[1] == "EQUIVALENT" or row[1] == "EQ":
                    eq_count += 1
                    eq_time += float(row[2])
                elif row[1] == "NOT EQUIVALENT" or row[1] == "NEQ":
                    neq_count += 1
                    neq_time += float(row[2])
                elif row[1] == "UNKNOWN" or row[1] == None:
                    unknown_count += 1
                    unk_time += float(row[2])
                elif row[1] == "MAYBE_EQ":
                    maybe_eq_count += 1
                    maybe_eq_time += float(row[2])
                elif row[1] == "MAYBE_NEQ":
                    maybe_neq_count += 1
                    maybe_neq_time += float(row[2])
                total_time += float(row[2])
                cur_row += 1
    outWorksheet.write("D2", eq_count)
    outWorksheet.write("E2", neq_count)
    outWorksheet.write("F2", unknown_count)
    outWorksheet.write("G2", total_time)
    outWorksheet.write("H2", total_time/(1000*(cur_row - 2)))
    outWorksheet.write("I2", eq_time/(1000*eq_count))
    outWorksheet.write("J2", neq_time/(1000*neq_count))
    outWorksheet.write("K2", unk_time/(1000*unknown_count))
    outWorksheet.write("L2", maybe_eq_count)
    outWorksheet.write("M2", maybe_neq_count)
    if maybe_eq_count!=0:
        outWorksheet.write("N2", maybe_eq_time/(1000*maybe_eq_count))
    if maybe_neq_count!=0:
        outWorksheet.write("O2", maybe_neq_time/(1000*maybe_neq_count))
    outWorkbook.close()
